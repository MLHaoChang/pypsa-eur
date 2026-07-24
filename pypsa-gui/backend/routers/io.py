from __future__ import annotations

import io
import pathlib
import tempfile
import zipfile

from fastapi import APIRouter, File, UploadFile
from models.schemas import ImportSummary
from services import change_log_service
from services.pypsa_service import PyPSAService
from services.upload_guard import read_capped, safe_extract
from starlette.responses import StreamingResponse

router = APIRouter()

BASE_DIR = pathlib.Path(__file__).parent.parent


def _build_summary(n) -> ImportSummary:
    return ImportSummary(
        buses=len(n.buses),
        generators=len(n.generators),
        lines=len(n.lines),
        links=len(n.links),
        storage_units=len(n.storage_units),
        stores=len(n.stores),
        loads=len(n.loads),
        transformers=len(n.transformers),
        snapshots=len(n.snapshots),
    )


# ── Exports ───────────────────────────────────────────────────────────────────

# Each export is split into a `_*_bytes()/_*_text()` generator (materialises the
# file content + audit-logs) and a thin route that wraps it in a StreamingResponse.
# The chat-tool layer calls the generator directly so it can persist the bytes as
# a downloadable `agent_export` artifact — a StreamingResponse can't be returned
# as a chat-tool result (json.dumps stringifies it to "<StreamingResponse ...>").


def _export_netcdf_bytes() -> bytes:
    n = PyPSAService.get_network()
    with tempfile.NamedTemporaryFile(suffix=".nc", delete=False) as f:
        tmp = pathlib.Path(f.name)
    with PyPSAService.get_netcdf_io_lock():
        n.export_to_netcdf(str(tmp))
    data = tmp.read_bytes()
    tmp.unlink(missing_ok=True)
    change_log_service.log("export", "Network", "network.nc", "Exported network as NetCDF (.nc)")
    return data


@router.get("/export/netcdf")
def export_netcdf():
    return StreamingResponse(
        io.BytesIO(_export_netcdf_bytes()),
        media_type="application/x-netcdf",
        headers={"Content-Disposition": "attachment; filename=network.nc"},
    )


def _export_csv_zip_bytes() -> bytes:
    n = PyPSAService.get_network()
    buf = io.BytesIO()
    with tempfile.TemporaryDirectory() as tmpdir:
        n.export_to_csv_folder(tmpdir)
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in pathlib.Path(tmpdir).rglob("*"):
                if f.is_file():
                    zf.write(f, f.relative_to(tmpdir))
    change_log_service.log("export", "Network", "network_csv.zip", "Exported network as CSV (zip)")
    return buf.getvalue()


@router.get("/export/csv")
def export_csv():
    return StreamingResponse(
        io.BytesIO(_export_csv_zip_bytes()),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=network_csv.zip"},
    )


def _export_excel_bytes() -> bytes:
    n = PyPSAService.get_network()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = pathlib.Path(f.name)
    try:
        n.export_to_excel(str(tmp))
        data = tmp.read_bytes()
    except AttributeError:
        # Fallback: write each component DataFrame as a sheet
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cname in ["buses", "carriers", "generators", "lines", "links",
                       "storage_units", "stores", "loads", "transformers"]:
            df = getattr(n, cname, None)
            if df is not None and not df.empty:
                ws = wb.create_sheet(cname[:31])
                rows = df.reset_index().values.tolist()
                ws.append(list(df.reset_index().columns))
                for row in rows:
                    ws.append(row)
        wb.save(str(tmp))
        data = tmp.read_bytes()
    finally:
        tmp.unlink(missing_ok=True)
    change_log_service.log("export", "Network", "network.xlsx", "Exported network as Excel (.xlsx)")
    return data


@router.get("/export/excel")
def export_excel():
    return StreamingResponse(
        io.BytesIO(_export_excel_bytes()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=network.xlsx"},
    )


def _export_matpower_text() -> str:
    from jinja2 import Environment, FileSystemLoader
    n = PyPSAService.get_network()
    bus_idx = {name: i + 1 for i, name in enumerate(n.buses.index)}
    control_map = {"Slack": 3, "PV": 2, "PQ": 1}

    buses = []
    for name, row in n.buses.iterrows():
        Pd = 0.0
        if not n.loads.empty:
            bus_loads = n.loads[n.loads.bus == name]
            Pd = float(bus_loads["p_set"].sum()) if not bus_loads.empty else 0.0
        buses.append({
            "bus_i": bus_idx[name],
            "type": control_map.get(getattr(row, "control", "PQ"), 1),
            "Pd": round(Pd, 4),
            "baseKV": float(row.get("v_nom", 1.0)),
        })

    gens = []
    for name, row in n.generators.iterrows():
        gens.append({
            "bus_i": bus_idx.get(row.bus, 1),
            "Pg": round(float(row.get("p_nom", 0)) * 0.5, 4),
            "Pmax": round(float(row.get("p_nom", 0)), 4),
            "Pmin": round(float(row.get("p_nom_min", 0)), 4),
            "marginal_cost": round(float(row.get("marginal_cost", 0)), 4),
        })

    branches = []
    for name, row in n.lines.iterrows():
        branches.append({
            "fbus": bus_idx.get(row.bus0, 1),
            "tbus": bus_idx.get(row.bus1, 2),
            "r": round(float(row.get("r", 0)), 6),
            "x": round(float(row.get("x", 0.01)), 6),
            "b": round(float(row.get("b", 0)), 6),
            "rateA": round(float(row.get("s_nom", 0)), 4),
        })

    env = Environment(loader=FileSystemLoader(str(BASE_DIR / "templates")))
    tmpl = env.get_template("matpower.jinja2")
    content = tmpl.render(
        name=n.name,
        baseMVA=100,
        buses=buses,
        gens=gens,
        branches=branches,
    )
    change_log_service.log("export", "Network", "network.m", "Exported network as MATPOWER (.m)")
    return content


@router.get("/export/matpower")
def export_matpower():
    return StreamingResponse(
        io.StringIO(_export_matpower_text()),
        media_type="text/plain",
        headers={"Content-Disposition": "attachment; filename=network.m"},
    )


# ── Imports ───────────────────────────────────────────────────────────────────

def _reset_with_ts_clear() -> None:
    """Reset network and clear any user-uploaded time series."""
    from routers.network import _restore_user_ts
    PyPSAService.reset_network()
    _restore_user_ts({})


@router.post("/import/netcdf")
async def import_netcdf(file: UploadFile = File(...)):
    data = await read_capped(file)
    with tempfile.NamedTemporaryFile(suffix=".nc", delete=False) as f:
        f.write(data)
        tmp = pathlib.Path(f.name)
    try:
        with PyPSAService.get_lock():
            _reset_with_ts_clear()
            n = PyPSAService.get_network()
            with PyPSAService.get_netcdf_io_lock():
                n.import_from_netcdf(str(tmp))
    finally:
        tmp.unlink(missing_ok=True)
    summary = _build_summary(n)
    change_log_service.log(
        "import", "Network", file.filename or "network.nc",
        f"Imported NetCDF '{file.filename}': {summary.buses} buses, {summary.generators} generators, {summary.snapshots} snapshots",
    )
    return summary


@router.post("/import/csv")
async def import_csv(file: UploadFile = File(...)):
    import zipfile as _zip
    data = await read_capped(file)
    with tempfile.TemporaryDirectory() as tmpdir:
        with _zip.ZipFile(io.BytesIO(data)) as zf:
            safe_extract(zf, tmpdir)
        with PyPSAService.get_lock():
            _reset_with_ts_clear()
            n = PyPSAService.get_network()
            n.import_from_csv_folder(tmpdir)
    summary = _build_summary(n)
    change_log_service.log(
        "import", "Network", file.filename or "network_csv.zip",
        f"Imported CSV zip '{file.filename}': {summary.buses} buses, {summary.generators} generators, {summary.snapshots} snapshots",
    )
    return summary


@router.post("/import/excel")
async def import_excel(file: UploadFile = File(...)):
    import openpyxl
    data = await read_capped(file)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    component_map = {
        "buses": "Bus", "generators": "Generator", "lines": "Line",
        "links": "Link", "loads": "Load", "storage_units": "StorageUnit",
        "stores": "Store", "transformers": "Transformer",
    }
    with PyPSAService.get_lock():
        _reset_with_ts_clear()
        n = PyPSAService.get_network()
        for sheet_name in wb.sheetnames:
            comp_class = component_map.get(sheet_name)
            if comp_class is None:
                continue
            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                name = row_dict.pop("name", row_dict.pop(headers[0], None))
                if name:
                    try:
                        n.add(comp_class, str(name), **{k: v for k, v in row_dict.items() if v is not None})
                    except Exception:
                        pass
    summary = _build_summary(n)
    change_log_service.log(
        "import", "Network", file.filename or "network.xlsx",
        f"Imported Excel '{file.filename}': {summary.buses} buses, {summary.generators} generators",
    )
    return summary


@router.post("/import/matpower")
async def import_matpower(file: UploadFile = File(...)):
    data = (await read_capped(file)).decode("utf-8")
    with PyPSAService.get_lock():
        _reset_with_ts_clear()
        n = PyPSAService.get_network()
        _parse_matpower(n, data)
    summary = _build_summary(n)
    change_log_service.log(
        "import", "Network", file.filename or "network.m",
        f"Imported MATPOWER '{file.filename}': {summary.buses} buses, {summary.generators} generators",
    )
    return summary


def _parse_matpower(n, text: str) -> None:
    import re


    def _section(label):
        pattern = rf"mpc\.{label}\s*=\s*\[(.*?)\]"
        m = re.search(pattern, text, re.DOTALL)
        if not m:
            return []
        rows = []
        for line in m.group(1).strip().splitlines():
            line = re.sub(r"%.*", "", line).strip().rstrip(";").strip()
            if line:
                rows.append([float(x) for x in line.split()])
        return rows

    buses_data = _section("bus")
    bus_map = {}
    for row in buses_data:
        bus_i = int(row[0])
        bus_name = f"bus_{bus_i}"
        bus_map[bus_i] = bus_name
        n.add("Bus", bus_name, v_nom=float(row[9]) if len(row) > 9 else 1.0)

    for i, row in enumerate(_section("gen")):
        bus_i = int(row[0])
        n.add("Generator", f"gen_{i+1}",
              bus=bus_map.get(bus_i, f"bus_{bus_i}"),
              p_nom=float(row[8]) if len(row) > 8 else 0.0,
              p_nom_min=float(row[9]) if len(row) > 9 else 0.0)

    for i, row in enumerate(_section("branch")):
        fbus, tbus = int(row[0]), int(row[1])
        n.add("Line", f"line_{i+1}",
              bus0=bus_map.get(fbus, f"bus_{fbus}"),
              bus1=bus_map.get(tbus, f"bus_{tbus}"),
              r=float(row[2]), x=float(row[3]),
              s_nom=float(row[5]) if len(row) > 5 else 0.0)
