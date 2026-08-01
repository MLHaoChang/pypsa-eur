"""The workbook: provenance, sheet layout and the two scopes."""
import io

import openpyxl

from tests.conftest import build_network

URL = "/api/results/asset/Generator/gas/export.xlsx"


def _book(resp):
    assert resp.status_code == 200
    return openpyxl.load_workbook(io.BytesIO(resp.content))


def test_configured_scope_writes_about_summary_and_the_category(
        client, install_network):
    install_network(build_network(solve=True))
    wb = _book(client.get(URL, params={
        "scope": "view", "category": "dispatch", "metrics": "p,energy_mwh"}))
    # "Key results" carries the headline KPIs lifted from every category. A
    # view-scope export builds it too — it is the page a reader opens first,
    # and it must not depend on which tab happened to be on screen.
    assert wb.sheetnames == ["About", "Key results", "Summary", "Dispatch"]


def test_about_sheet_carries_every_provenance_field(client, install_network):
    install_network(build_network(solve=True))
    wb = _book(client.get(URL, params={"scope": "view", "category": "dispatch",
                                       "metrics": "p"}))
    keys = {row[0] for row in wb["About"].iter_rows(min_col=1, max_col=1,
                                                    values_only=True) if row[0]}
    for expected in ("Asset", "Component class", "Category", "View mode",
                     "Result source", "Horizon from", "Horizon to", "Period",
                     "Objective", "PyPSA version", "Generated at"):
        assert expected in keys, f"About sheet is missing '{expected}'"


def test_data_sheet_header_matches_the_columns_contract(client, install_network):
    install_network(build_network(solve=True))
    wb = _book(client.get(URL, params={"scope": "view", "category": "dispatch",
                                       "metrics": "p"}))
    header = [c.value for c in wb["Dispatch"][1]]
    assert header[0] == "snapshot"
    assert "Active power (MW)" in header


def test_duration_mode_writes_rank_and_percentile_columns(client, install_network):
    install_network(build_network(solve=True))
    wb = _book(client.get(URL, params={"scope": "view", "category": "dispatch",
                                       "metrics": "p", "mode": "duration"}))
    header = [c.value for c in wb["Dispatch"][1]]
    assert header[0] == "rank"
    assert header[1] == "pct_of_hours"


def test_full_scope_writes_every_applicable_category(client, install_network):
    install_network(build_network(solve=True))
    wb = _book(client.get(URL, params={"scope": "full"}))
    assert "Dispatch" in wb.sheetnames
    assert "Capacity" in wb.sheetnames
    assert "Storage" not in wb.sheetnames, "n/a categories must be omitted"
    about = {row[0]: row[1] for row in
             wb["About"].iter_rows(min_col=1, max_col=2, values_only=True)}
    assert about["Category"] == "(all applicable)", (
        "a full export spans several categories — naming it after whichever "
        "one happened to run first (always Summary) is misleading")


def test_full_scope_omits_a_category_whose_build_response_raises(
        client, install_network, monkeypatch):
    """
    A single category raising must cost that category, not the workbook.

    Monkeypatches `build_response` so exactly ONE category (prices — normally
    `ok` for a solved Generator, per the existing full-scope test) raises.
    Without the try/except in `export.build_workbook`, this request 500s and
    `_book` fails before any assertion below runs — so this test only passes
    if the except branch actually caught the exception.
    """
    install_network(build_network(solve=True))
    from services.asset_results import export as xls

    real_build_response = xls.build_response

    def flaky(n, component_class, name, *, category, **kwargs):
        if category == "prices":
            raise RuntimeError("boom")
        return real_build_response(
            n, component_class, name, category=category, **kwargs)

    monkeypatch.setattr(xls, "build_response", flaky)

    wb = _book(client.get(URL, params={"scope": "full"}))
    assert "Dispatch" in wb.sheetnames
    assert "Capacity" in wb.sheetnames
    assert "Prices & duals" not in wb.sheetnames, (
        "the category whose build_response raised must not get a sheet")
    text = "\n".join(
        str(row[0]) + "|" + str(row[1])
        for row in wb["About"].iter_rows(min_col=1, max_col=2, values_only=True)
    )
    assert "Omitted: Prices & duals|failed to compute: boom" in text


def test_full_scope_lists_the_omitted_categories_in_about(client, install_network):
    install_network(build_network(solve=True))
    wb = _book(client.get(URL, params={"scope": "full"}))
    text = "\n".join(
        str(row[0]) + "|" + str(row[1])
        for row in wb["About"].iter_rows(min_col=1, max_col=2, values_only=True)
    )
    assert "Storage" in text and "store energy" in text


def test_unsolved_network_still_exports_the_summary(client, install_network):
    install_network(build_network(solve=False))
    wb = _book(client.get(URL, params={"scope": "full"}))
    assert "Summary" in wb.sheetnames
    assert "Dispatch" not in wb.sheetnames


def test_view_scope_returns_a_clean_500_when_its_only_category_raises(
        client, install_network, monkeypatch):
    """
    `scope=view` runs exactly ONE category. If that raises, `first_resp`
    never gets set — before this fix the only guard was a bare `assert`,
    which is an unhandled AssertionError (opaque 500, no detail) and is
    compiled out entirely under `python -O`. The export route has no
    try/except of its own (unlike the sibling read endpoint), so this must
    be an explicit, clean error instead.
    """
    install_network(build_network(solve=True))
    from services.asset_results import export as xls

    def boom(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr(xls, "build_response", boom)
    resp = client.get(URL, params={"scope": "view", "category": "dispatch",
                                   "metrics": "p"})
    assert resp.status_code == 500
    assert "detail" in resp.json()


def test_bad_scope_is_422(client, install_network):
    install_network(build_network(solve=True))
    assert client.get(URL, params={"scope": "nope"}).status_code == 422


def test_bad_mode_is_422(client, install_network):
    install_network(build_network(solve=True))
    resp = client.get(URL, params={"scope": "view", "category": "dispatch",
                                   "metrics": "p", "mode": "nope"})
    assert resp.status_code == 422
