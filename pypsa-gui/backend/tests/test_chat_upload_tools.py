"""
Phase B — chat tools that consume + produce uploads.

Covers each tool's happy path + the structured `error_kind` cases for
apply_demand_from_excel + the export tools' sanitation/cap guards.
"""
from __future__ import annotations

import base64
import io

import pandas as pd
import pytest
from fastapi import HTTPException
from openpyxl import Workbook

from services import chat_tools, upload_service
from services.pypsa_service import PyPSAService

from tests.conftest import build_network


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ─────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────


@pytest.fixture
def install_with_uploads(tmp_projects_dir, install_network):
    """Install a network bound to project 'P' + create the project dir on disk
    so upload_service.add_upload has a real uploads/ to write into."""
    n = build_network()
    install_network(n, name="P")
    (tmp_projects_dir / "P").mkdir(parents=True, exist_ok=True)
    (tmp_projects_dir / "P" / "network.nc").write_bytes(b"")  # marker file
    return tmp_projects_dir / "P"


def _make_xlsx(rows: list[list], headers: list[str] | None = None,
               sheet: str = "Sheet1") -> bytes:
    """Return the bytes of a tiny xlsx workbook with `rows`."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────
# list_uploads / read_upload_meta / delete_upload
# ─────────────────────────────────────────────────────────────────────────


class TestListUploads:
    def test_empty_when_no_uploads(self, install_with_uploads):
        assert chat_tools.list_uploads() == []

    def test_returns_chip_shape(self, install_with_uploads):
        upload_service.add_upload("P", b"hello,world\n", "test.csv", "text/csv")
        items = chat_tools.list_uploads()
        assert len(items) == 1
        assert set(items[0].keys()) == {
            "file_id", "filename", "mime", "kind", "size_kb",
            "uploaded_at", "page_count",
        }
        assert items[0]["filename"] == "test.csv"
        assert items[0]["kind"] == "user_upload"


class TestNoActiveProject:
    def test_no_active_project_raises_400(self, tmp_projects_dir, monkeypatch):
        """All upload tools must refuse when no project is loaded."""
        # tmp_projects_dir is the fixture that sets PROJECTS_DIR; we DON'T
        # call install_network so loaded_project remains None.
        PyPSAService.set_loaded_project(None)
        with pytest.raises(HTTPException) as ei:
            chat_tools.list_uploads()
        assert ei.value.status_code == 400
        assert ei.value.detail["error_kind"] == "no_active_project"


class TestDeleteUpload:
    def test_delete_existing(self, install_with_uploads):
        meta = upload_service.add_upload("P", b"a,b\n1,2\n", "x.csv", "text/csv")
        resp = chat_tools.delete_upload(meta.file_id)
        assert resp == {"deleted": True, "file_id": meta.file_id}
        assert chat_tools.list_uploads() == []

    def test_delete_missing(self, install_with_uploads):
        resp = chat_tools.delete_upload("deadbeef00000000")
        assert resp["deleted"] is False
        assert resp["reason"] == "not_found"


# ─────────────────────────────────────────────────────────────────────────
# read_excel_sheet
# ─────────────────────────────────────────────────────────────────────────


class TestReadExcelSheet:
    def test_csv_returns_shape(self, install_with_uploads):
        meta = upload_service.add_upload(
            "P", b"col1,col2\n1,2\n3,4\n5,6\n", "data.csv", "text/csv",
        )
        out = chat_tools.read_excel_sheet(meta.file_id)
        assert out["columns"] == ["col1", "col2"]
        assert out["total_rows"] == 3
        assert out["total_cols"] == 2
        assert out["rows"] == [[1, 2], [3, 4], [5, 6]]
        assert out["truncated"] is False

    def test_max_rows_truncates(self, install_with_uploads):
        csv_payload = b"x\n" + b"\n".join(str(i).encode() for i in range(500))
        meta = upload_service.add_upload("P", csv_payload, "big.csv", "text/csv")
        out = chat_tools.read_excel_sheet(meta.file_id, max_rows=100)
        assert out["total_rows"] == 500
        assert len(out["rows"]) == 100
        assert out["truncated"] is True

    def test_xlsx_first_sheet(self, install_with_uploads):
        payload = _make_xlsx([[1, 2], [3, 4]], headers=["a", "b"])
        meta = upload_service.add_upload("P", payload, "data.xlsx", _XLSX_MIME)
        out = chat_tools.read_excel_sheet(meta.file_id)
        assert out["columns"] == ["a", "b"]
        assert out["total_rows"] == 2

    def test_read_excel_returns_real_sheet_name_not_placeholder(
        self, install_with_uploads,
    ):
        """
        Regression for the 2026-06-08 incident: `read_excel_sheet` used to
        return `sheet_name: "(first sheet)"` when called without a
        sheet_name arg. The agent then reused that placeholder string in
        the next `apply_demand_from_excel` call, pandas couldn't find a
        sheet named literally `(first sheet)`, and the user saw
        `excel_parse_failed`. The response must return the ACTUAL pandas-
        resolved sheet name (e.g. "Sheet1") so a round-trip works.
        """
        # _make_xlsx names the sheet "Sheet1" — the openpyxl default.
        payload = _make_xlsx([[1, 2]], headers=["a", "b"], sheet="Sheet1")
        meta = upload_service.add_upload("P", payload, "x.xlsx", _XLSX_MIME)
        out = chat_tools.read_excel_sheet(meta.file_id)  # no sheet_name
        assert out["sheet_name"] == "Sheet1"  # real, not "(first sheet)"
        assert out["available_sheets"] == ["Sheet1"]
        # Round-trip: feeding the returned name back must work
        out2 = chat_tools.read_excel_sheet(
            meta.file_id, sheet_name=out["sheet_name"],
        )
        assert out2["sheet_name"] == "Sheet1"

    def test_read_excel_exposes_all_sheet_names(self, install_with_uploads):
        """
        Multi-sheet workbook: `available_sheets` lists every sheet so the
        agent doesn't have to guess. Without this the agent would have to
        retry random names ("Sheet1", "Sheet2", …) on every multi-sheet
        upload.
        """
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = "Demand"
        wb.active.append(["t", "v"])
        wb.active.append([1, 100])
        wb.create_sheet("Solar")
        wb["Solar"].append(["t", "irr"])
        wb["Solar"].append([1, 250])
        buf = io.BytesIO()
        wb.save(buf)
        meta = upload_service.add_upload(
            "P", buf.getvalue(), "multi.xlsx", _XLSX_MIME,
        )
        out = chat_tools.read_excel_sheet(meta.file_id)
        # First sheet is the active one — "Demand"
        assert out["sheet_name"] == "Demand"
        assert out["available_sheets"] == ["Demand", "Solar"]
        # Reading the named second sheet works
        out2 = chat_tools.read_excel_sheet(meta.file_id, sheet_name="Solar")
        assert out2["sheet_name"] == "Solar"
        assert out2["columns"] == ["t", "irr"]

    def test_read_excel_unknown_sheet_name_rejected(self, install_with_uploads):
        """
        Asking for a non-existent sheet name returns a clean
        `excel_parse_failed` with the available list in the message —
        actionable error rather than a cryptic pandas trace.
        """
        payload = _make_xlsx([[1, 2]], headers=["a", "b"], sheet="Real")
        meta = upload_service.add_upload("P", payload, "x.xlsx", _XLSX_MIME)
        with pytest.raises(HTTPException) as ei:
            chat_tools.read_excel_sheet(meta.file_id, sheet_name="NotReal")
        assert ei.value.detail["error_kind"] == "excel_parse_failed"
        assert "Real" in ei.value.detail["message"]  # available listed


# ─────────────────────────────────────────────────────────────────────────
# apply_demand_from_excel
# ─────────────────────────────────────────────────────────────────────────


class TestApplyDemandFromExcel:
    def _make_aligned_csv(self, snapshots, values: list[float]) -> bytes:
        rows = []
        for sn, val in zip(snapshots, values):
            rows.append(f"{sn},{val}")
        return b"timestamp,demand\n" + "\n".join(rows).encode() + b"\n"

    def test_happy_path_writes_p_set(self, install_with_uploads):
        n = PyPSAService.get_network()
        # build_network creates a 4-snapshot network with one Load 'L1'
        payload = self._make_aligned_csv(
            n.snapshots, [10.0, 20.0, 30.0, 40.0],
        )
        meta = upload_service.add_upload(
            "P", payload, "demand.csv", "text/csv",
        )
        result = chat_tools.apply_demand_from_excel(
            file_id=meta.file_id,
            sheet_name=None,
            time_col="timestamp",
            value_col="demand",
            load_name="L1",
        )
        assert result["applied"] is True
        assert result["rows"] == 4
        assert result["min"] == 10.0
        assert result["max"] == 40.0
        # Verify the dataframe write took effect
        assert list(n.loads_t.p_set["L1"]) == [10.0, 20.0, 30.0, 40.0]

    def test_unknown_load_raises_load_not_found(self, install_with_uploads):
        meta = upload_service.add_upload("P", b"a,b\n1,2\n", "x.csv", "text/csv")
        with pytest.raises(HTTPException) as ei:
            chat_tools.apply_demand_from_excel(
                file_id=meta.file_id, sheet_name=None,
                time_col="a", value_col="b", load_name="DOES_NOT_EXIST",
            )
        assert ei.value.status_code == 400
        assert ei.value.detail["error_kind"] == "load_not_found"

    def test_missing_time_column_raises(self, install_with_uploads):
        meta = upload_service.add_upload(
            "P", b"x,y\n1,2\n", "x.csv", "text/csv",
        )
        with pytest.raises(HTTPException) as ei:
            chat_tools.apply_demand_from_excel(
                file_id=meta.file_id, sheet_name=None,
                time_col="NOT_A_COLUMN", value_col="y", load_name="L1",
            )
        assert ei.value.detail["error_kind"] == "time_column_parse_error"

    def test_missing_value_column_raises(self, install_with_uploads):
        meta = upload_service.add_upload(
            "P", b"time,y\n2025-01-01,2\n", "x.csv", "text/csv",
        )
        with pytest.raises(HTTPException) as ei:
            chat_tools.apply_demand_from_excel(
                file_id=meta.file_id, sheet_name=None,
                time_col="time", value_col="NOT_A_COLUMN", load_name="L1",
            )
        assert ei.value.detail["error_kind"] == "value_column_parse_error"

    def test_non_numeric_value_raises(self, install_with_uploads):
        n = PyPSAService.get_network()
        ts = [str(s) for s in n.snapshots]
        rows = "\n".join(f"{ts[i]},text_{i}" for i in range(len(ts)))
        meta = upload_service.add_upload(
            "P", f"time,val\n{rows}\n".encode(), "x.csv", "text/csv",
        )
        with pytest.raises(HTTPException) as ei:
            chat_tools.apply_demand_from_excel(
                file_id=meta.file_id, sheet_name=None,
                time_col="time", value_col="val", load_name="L1",
            )
        assert ei.value.detail["error_kind"] == "value_column_parse_error"

    def test_snapshot_count_mismatch(self, install_with_uploads):
        n = PyPSAService.get_network()
        # Only 2 rows; network has 4 snapshots
        payload = b"time,demand\n2025-01-01,10\n2025-01-02,20\n"
        meta = upload_service.add_upload("P", payload, "x.csv", "text/csv")
        with pytest.raises(HTTPException) as ei:
            chat_tools.apply_demand_from_excel(
                file_id=meta.file_id, sheet_name=None,
                time_col="time", value_col="demand", load_name="L1",
            )
        assert ei.value.detail["error_kind"] == "snapshot_count_mismatch"

    def test_apply_demand_works_when_sheet_name_omitted(
        self, install_with_uploads,
    ):
        """
        Regression for the 2026-06-08 incident: the agent called
        `apply_demand_from_excel(file_id=..., time_col=..., value_col=...,
        load_name=...)` without `sheet_name` because the schema lists it as
        optional — but the Python function signature had it as required
        positional, so the dispatcher raised `TypeError: missing 1 required
        positional argument`. The fix moves `sheet_name` to a keyword arg
        with `None` default so omitting it works AND the call defaults to
        the first sheet.

        Also asserts the placeholder-string round-trip from the old
        `read_excel_sheet` is handled gracefully — feeding `(first sheet)`
        as sheet_name (the legacy placeholder) is treated the same as
        `None` instead of raising.
        """
        n = PyPSAService.get_network()
        # Use a CSV here so the test stays focused on the kwarg-default
        # bug rather than xlsx parsing.
        payload = self._make_aligned_csv(
            n.snapshots, [10.0, 20.0, 30.0, 40.0],
        )
        meta = upload_service.add_upload("P", payload, "demand.csv", "text/csv")

        # Call WITHOUT sheet_name — must not raise TypeError.
        result = chat_tools.apply_demand_from_excel(
            file_id=meta.file_id,
            time_col="timestamp",
            value_col="demand",
            load_name="L1",
        )
        assert result["applied"] is True

        # Feed the legacy placeholder string back — must be treated as
        # "default sheet" not as a literal sheet name to look up.
        result2 = chat_tools.apply_demand_from_excel(
            file_id=meta.file_id,
            time_col="timestamp",
            value_col="demand",
            load_name="L1",
            sheet_name="(first sheet)",
        )
        assert result2["applied"] is True

    def test_multi_period_auto_tile_one_year_profile(self, install_with_uploads):
        """
        A 1-year operational profile is the typical multi-period demand
        input — the user expects the same hourly shape to repeat each
        investment period. The agent must NOT have to manually tile the
        Excel rows 3× to match a 3-period network.

        Setup: 3 investment periods × 4 operational hours = 12 snapshots.
        Excel: 4 rows (one operational year). Expected: auto-tile to 12
        values, returning `auto_tiled=True, tile_factor=3` so the agent
        can announce the broadcast in its reply.

        Regression for the live 2026-06-08 incident: agent had to hand-
        build a 26,280-row CSV to apply an 8,760-row 1-year profile to a
        3-period network. Now it just works in one tool call.
        """
        import pandas as pd
        import pypsa

        n_mp = pypsa.Network()
        op = pd.date_range("2025-01-01", periods=4, freq="h")
        mi = pd.MultiIndex.from_product(
            [[2026, 2027, 2028], op], names=["period", "timestep"],
        )
        mi.name = "snapshot"
        n_mp.set_snapshots(mi)
        n_mp.investment_periods = [2026, 2027, 2028]
        n_mp.add("Bus", "B1")
        n_mp.add("Load", "L1", bus="B1", p_set=0.0)
        PyPSAService.set_network(n_mp)
        PyPSAService.set_loaded_project("P")

        # 4 rows = one operational year. The network has 12 snapshots
        # (3 periods × 4 hours). Auto-tile should fire.
        payload = (
            b"time,demand\n"
            b"2025-01-01 00:00,100\n"
            b"2025-01-01 01:00,200\n"
            b"2025-01-01 02:00,300\n"
            b"2025-01-01 03:00,400\n"
        )
        meta = upload_service.add_upload("P", payload, "year.csv", "text/csv")

        result = chat_tools.apply_demand_from_excel(
            file_id=meta.file_id, sheet_name=None,
            time_col="time", value_col="demand", load_name="L1",
        )
        assert result["applied"] is True
        assert result["rows"] == 12  # tiled to full snapshot count
        assert result["auto_tiled"] is True
        assert result["tile_factor"] == 3

        live = PyPSAService.get_network()
        # Values tile cleanly across all 3 periods: [100, 200, 300, 400] × 3
        expected = [100, 200, 300, 400, 100, 200, 300, 400, 100, 200, 300, 400]
        assert list(live.loads_t.p_set["L1"]) == expected

    def test_multi_period_non_divisible_count_rejected(self, install_with_uploads):
        """
        A 6-row profile against a 3-period × 4-hour (= 12 snapshots) network
        is neither a full match (12) nor an exact per-period match (4); we
        refuse rather than guess what the user meant.
        """
        import pandas as pd
        import pypsa

        n_mp = pypsa.Network()
        op = pd.date_range("2025-01-01", periods=4, freq="h")
        mi = pd.MultiIndex.from_product(
            [[2026, 2027, 2028], op], names=["period", "timestep"],
        )
        mi.name = "snapshot"
        n_mp.set_snapshots(mi)
        n_mp.investment_periods = [2026, 2027, 2028]
        n_mp.add("Bus", "B1")
        n_mp.add("Load", "L1", bus="B1", p_set=0.0)
        PyPSAService.set_network(n_mp)
        PyPSAService.set_loaded_project("P")

        # 6 rows — doesn't match 12 (full) or 4 (per-period)
        payload = (
            b"time,demand\n"
            b"2025-01-01 00:00,10\n2025-01-01 01:00,20\n"
            b"2025-01-01 02:00,30\n2025-01-01 03:00,40\n"
            b"2025-01-01 04:00,50\n2025-01-01 05:00,60\n"
        )
        meta = upload_service.add_upload("P", payload, "weird.csv", "text/csv")

        with pytest.raises(HTTPException) as ei:
            chat_tools.apply_demand_from_excel(
                file_id=meta.file_id, sheet_name=None,
                time_col="time", value_col="demand", load_name="L1",
            )
        assert ei.value.detail["error_kind"] == "snapshot_count_mismatch"
        # Error message must explain BOTH valid row counts so the agent
        # can guide the user.
        msg = ei.value.detail["message"]
        assert "12" in msg  # full coverage
        assert "4" in msg   # per-period


    def test_multi_period_positional_alignment(self, install_with_uploads):
        """
        Multi-period network (e.g. investment periods 2026/2027/2028 with
        hourly operational year): the user's xlsx has flat timestamps but
        the row count matches len(n.snapshots). `apply_demand_from_excel`
        used to block these with `snapshot_range_mismatch` because the
        flat ISO timestamps couldn't equal the MultiIndex tuples. The fix
        aligns POSITIONALLY when the network is multi-period — the strict
        range check only applies to flat networks.

        Regression for the live 2026-06-08 incident on a 3-year
        (26,280-snapshot) multi-period network with a matching xlsx.
        """
        import pandas as pd
        import pypsa

        # Build a tiny multi-period network: 2 periods × 3 operational hours
        n_mp = pypsa.Network()
        op = pd.date_range("2025-01-01", periods=3, freq="h")
        mi = pd.MultiIndex.from_product([[2026, 2027], op], names=["period", "timestep"])
        mi.name = "snapshot"
        n_mp.set_snapshots(mi)
        n_mp.investment_periods = [2026, 2027]
        n_mp.add("Bus", "B1")
        n_mp.add("Load", "L1", bus="B1", p_set=0.0)
        PyPSAService.set_network(n_mp)
        PyPSAService.set_loaded_project("P")

        # 6 rows matching the 6-snapshot multi-period network. The flat
        # timestamps don't match the MultiIndex tuples — that's the whole
        # point of the test.
        payload = (
            b"time,demand\n"
            b"2025-01-01 00:00,10\n"
            b"2025-01-01 01:00,20\n"
            b"2025-01-01 02:00,30\n"
            b"2025-01-01 03:00,40\n"
            b"2025-01-01 04:00,50\n"
            b"2025-01-01 05:00,60\n"
        )
        meta = upload_service.add_upload("P", payload, "mp.csv", "text/csv")

        result = chat_tools.apply_demand_from_excel(
            file_id=meta.file_id, sheet_name=None,
            time_col="time", value_col="demand", load_name="L1",
        )
        assert result["applied"] is True
        assert result["rows"] == 6
        # Values landed at the right positions across both periods
        live = PyPSAService.get_network()
        assert list(live.loads_t.p_set["L1"]) == [10, 20, 30, 40, 50, 60]


# ─────────────────────────────────────────────────────────────────────────
# reconstruct_network_from_image — REAL impl in Phase C
#
# The phase_c_pending stub from Phase B is gone; the full vision-driven
# implementation lives in test_chat_multimodal.py::TestReconstructNetworkFromImage
# (those tests inject a fake Anthropic client). Phase B's wiring is now
# verified by the dispatcher-symmetry check at the bottom of this file.
# ─────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────
# Export tools (produce)
# ─────────────────────────────────────────────────────────────────────────


class TestExportToExcel:
    def test_writes_multi_sheet_workbook(self, install_with_uploads):
        meta = chat_tools.export_to_excel(
            sheets={
                "A": [["col1", "col2"], [1, 2], [3, 4]],
                "B": [["x"], [10], [20], [30]],
            },
            filename="report.xlsx",
        )
        assert meta["filename"] == "report.xlsx"
        assert meta["kind"] == "agent_export"
        # Round-trip: load the file_id back as an xlsx
        path = upload_service.get_upload_path("P", meta["file_id"])
        wb_back = pd.ExcelFile(path)
        assert set(wb_back.sheet_names) == {"A", "B"}
        df_a = wb_back.parse("A")
        assert list(df_a.columns) == ["col1", "col2"]
        assert df_a["col1"].tolist() == [1, 3]

    def test_filename_sanitised(self, install_with_uploads):
        meta = chat_tools.export_to_excel(
            sheets={"S": [["x"], [1]]},
            filename="../../etc/passwd.xlsx",
        )
        assert "/" not in meta["filename"]
        assert ".." not in meta["filename"]


class TestExportToCsv:
    def test_writes_rfc4180(self, install_with_uploads):
        meta = chat_tools.export_to_csv(
            rows=[[1, "a"], [2, "b"]],
            columns=["n", "letter"],
            filename="out.csv",
        )
        path = upload_service.get_upload_path("P", meta["file_id"])
        # Read RAW bytes — Python's text mode normalises \r\n → \n on
        # Windows under universal-newlines, hiding the RFC4180 line ending.
        raw = path.read_bytes()
        assert b"\r\n" in raw
        # Header preserved on the first physical line
        assert raw.split(b"\r\n")[0] == b"n,letter"

    def test_filename_sanitised(self, install_with_uploads):
        meta = chat_tools.export_to_csv(
            rows=[[1]], columns=["x"], filename="con.csv",
        )
        # CON is reserved on Windows — should get the underscore prefix
        assert meta["filename"].startswith("_")


class TestExportPreviewPng:
    def test_valid_png_succeeds(self, install_with_uploads):
        # Smallest valid PNG: 8-byte signature + IHDR chunk + IDAT + IEND
        # https://en.wikipedia.org/wiki/Portable_Network_Graphics
        # For our test, we just need the first 8 bytes to be PNG magic.
        valid_png = (
            b"\x89PNG\r\n\x1a\n"
            b"\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
            b"\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc\xf8\xcf\xc0\x00\x00\x00\x03\x00\x01"
            b"\xf3\xff\xa9\xed\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        meta = chat_tools.export_preview_png(
            filename="preview.png",
            png_bytes_b64=base64.b64encode(valid_png).decode(),
        )
        assert meta["filename"] == "preview.png"
        assert meta["mime"] == "image/png"

    def test_invalid_png_magic_rejected(self, install_with_uploads):
        # JPEG bytes; doesn't match PNG magic
        with pytest.raises(HTTPException) as ei:
            chat_tools.export_preview_png(
                filename="x.png",
                png_bytes_b64=base64.b64encode(b"\xff\xd8\xff").decode(),
            )
        assert ei.value.detail["error_kind"] == "invalid_png"

    def test_invalid_base64_rejected(self, install_with_uploads):
        with pytest.raises(HTTPException) as ei:
            chat_tools.export_preview_png(
                filename="x.png", png_bytes_b64="not!valid!base64!!!",
            )
        assert ei.value.detail["error_kind"] == "invalid_base64"


class TestExportChatSummary:
    def test_format_md(self, install_with_uploads):
        meta = chat_tools.export_chat_summary(format="md")
        assert meta["filename"].startswith("chat_summary_")
        assert meta["filename"].endswith(".md")
        assert meta["mime"] == "text/markdown"

    def test_format_txt(self, install_with_uploads):
        meta = chat_tools.export_chat_summary(format="txt")
        assert meta["filename"].endswith(".txt")
        assert meta["mime"] == "text/plain"

    def test_format_unsupported(self, install_with_uploads):
        with pytest.raises(HTTPException) as ei:
            chat_tools.export_chat_summary(format="pdf")
        assert ei.value.detail["error_kind"] == "format_not_supported"

    def test_custom_filename(self, install_with_uploads):
        meta = chat_tools.export_chat_summary(format="md", filename="my-export.md")
        assert meta["filename"] == "my-export.md"


# ─────────────────────────────────────────────────────────────────────────
# Schema / dispatcher invariants — Phase B raises tool count to 110
# ─────────────────────────────────────────────────────────────────────────


def test_phase_b_tool_count():
    from services.chat_tools_schema import TOOLS
    from services.chat_tools import DISPATCHERS
    assert len(TOOLS) == len(DISPATCHERS)
    # The hardcoded `assert len(TOOLS) == 111` that used to sit here went
    # stale the moment get_aggregate_load was added (112), and nobody noticed
    # because this file could not be collected — importing chat_tools raised
    # NameError on the unimplemented synthesis tools. It also contradicted
    # chat_tools_schema's own stated convention (v4-NIT-1 / v6-F4): "len(TOOLS)
    # is the only source of truth for the tool count — no pre-stated magic
    # number anywhere."
    #
    # The invariant above is the real check. Per-tool presence is covered by
    # test_phase_b_tools_have_callable_dispatchers below, which names the ten
    # Phase B tools explicitly instead of counting them.


def test_phase_b_tools_have_callable_dispatchers():
    from services.chat_tools import DISPATCHERS
    NEW = [
        "list_uploads", "read_upload_meta", "read_excel_sheet",
        "apply_demand_from_excel", "delete_upload",
        "reconstruct_network_from_image",
        "export_to_excel", "export_to_csv",
        "export_preview_png", "export_chat_summary",
    ]
    for name in NEW:
        assert name in DISPATCHERS, f"{name} missing from DISPATCHERS"
        assert callable(DISPATCHERS[name])
